"""
Export service for generating Excel and CSV attendance reports.

Exports are saved to the configured exports directory (Docker volume).
"""

import csv
import io
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import AttendanceRecord
from app.config import get_settings
from app.utils.logging import get_logger

settings = get_settings()
logger = get_logger()


class ExportService:
    def __init__(self):
        self.exports_dir = settings.EXPORTS_DIR

    def export_to_excel(
        self,
        records: List[AttendanceRecord],
        start_date: str,
        end_date: str,
        save_to_disk: bool = True
    ) -> tuple[bytes, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells('A1:G1')
        ws['A1'] = f"Attendance Report: {start_date} to {end_date}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:G2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal="center")

        headers = ["Name", "ID Number", "Date", "Time In", "Time Out", "Status", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, record in enumerate(records, 5):
            data = [
                record.name or "Unknown",
                record.employee_id or "N/A",
                record.date,
                record.time_in.strftime("%H:%M:%S") if record.time_in else "",
                record.time_out.strftime("%H:%M:%S") if record.time_out else "",
                record.status,
                record.notes or ""
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left" if col in [1, 7] else "center")

        column_widths = [25, 15, 12, 12, 12, 12, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        summary_row = len(records) + 6
        ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value=f"Total Records: {len(records)}")

        present_count = sum(1 for r in records if r.status == "present" or r.status == "completed")
        ws.cell(row=summary_row + 2, column=1, value=f"Present: {present_count}")

        unrecognized_count = sum(1 for r in records if not r.is_recognized)
        ws.cell(row=summary_row + 3, column=1, value=f"Unrecognized: {unrecognized_count}")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_bytes = output.getvalue()

        filename = f"attendance_{start_date}_to_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        if save_to_disk:
            self._save_to_disk(file_bytes, filename)

        return file_bytes, filename

    def export_to_csv(
        self,
        records: List[AttendanceRecord],
        start_date: str,
        end_date: str,
        save_to_disk: bool = True
    ) -> tuple[bytes, str]:
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["Name", "ID Number", "Date", "Time In", "Time Out", "Status", "Notes"])

        for record in records:
            writer.writerow([
                record.name or "Unknown",
                record.employee_id or "N/A",
                record.date,
                record.time_in.strftime("%H:%M:%S") if record.time_in else "",
                record.time_out.strftime("%H:%M:%S") if record.time_out else "",
                record.status,
                record.notes or ""
            ])

        file_bytes = output.getvalue().encode('utf-8')
        filename = f"attendance_{start_date}_to_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        if save_to_disk:
            self._save_to_disk(file_bytes, filename)

        return file_bytes, filename

    def _save_to_disk(self, file_bytes: bytes, filename: str) -> Path:
        self.exports_dir.mkdir(parents=True, exist_ok=True)
        file_path = self.exports_dir / filename

        with open(file_path, 'wb') as f:
            f.write(file_bytes)

        logger.info(f"Export saved to: {file_path}")
        return file_path

    def list_exports(self) -> List[dict]:
        if not self.exports_dir.exists():
            return []

        exports = []
        for file_path in self.exports_dir.iterdir():
            if file_path.suffix in ['.xlsx', '.csv']:
                stat = file_path.stat()
                exports.append({
                    'filename': file_path.name,
                    'size': stat.st_size,
                    'created': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    'format': file_path.suffix[1:]
                })

        return sorted(exports, key=lambda x: x['created'], reverse=True)

    def get_export_file(self, filename: str) -> Optional[Path]:
        file_path = self.exports_dir / filename
        if file_path.exists() and file_path.is_file():
            return file_path
        return None

    def delete_export(self, filename: str) -> bool:
        file_path = self.exports_dir / filename
        if file_path.exists() and file_path.is_file():
            file_path.unlink()
            logger.info(f"Export deleted: {filename}")
            return True
        return False


export_service = ExportService()


def get_export_service() -> ExportService:
    return export_service
